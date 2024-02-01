import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import openpyxl
import cx_Oracle

con = cx_Oracle.connect("lucle", "lucle", "192.168.0.118:1521/xe", encoding="UTF-8")
cursor = con.cursor()

def initialize_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option("detach", True)
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    return webdriver.Chrome(options=options)

def search_and_scrape(driver, search_query):
    data = []  # data 변수 정의
    total_price = 0  # 총 가격 합 변수 정의
    count = 0  # 상품 개수 카운트 변수 정의

    try:
        driver.get('https://www.11st.co.kr/')

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tSearch"]/form/fieldset/input')))
        search_input = driver.find_element(By.XPATH, '//*[@id="tSearch"]/form/fieldset/input')

        # 검색어를 함수 파라미터로 받아 사용
        ActionChains(driver).send_keys_to_element(search_input, search_query).perform()
        ActionChains(driver).send_keys(Keys.ENTER).perform()

        # 결과가 로딩될 때까지 대기
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, 'c-card-item')))
        time.sleep(2)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '11번가 상품 정보'

        current_page = 1
        max_pages = 3  # 가져올 최대 페이지 수를 설정합니다.

        while current_page <= max_pages:
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            product_items = driver.find_elements(By.CSS_SELECTOR, '#section_commonPrd .c-card-item')

            for item in product_items:
                try:
                    # 명시적인 대기 시간 추가
                    WebDriverWait(item, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'c-card-item__name')))
                    pName = item.find_element(By.CLASS_NAME, 'c-card-item__name').text.strip('상품명\n')


                    pPrice = item.find_element(By.CLASS_NAME, 'c-card-item__price').text.strip('~')
                    pPrice = pPrice.replace(',', '').strip('원')

                    
                    pDelivery_fee = ''
                    pDelivery_date = ''
                    pDelivery = item.find_element(By.CLASS_NAME, 'c-card-item__delivery').text

                    if pDelivery != '배송':
                        pDelivery = pDelivery.split('\n')
                        pDelivery_fee = pDelivery[1].strip('배송비 ')
                        pDelivery_date = '' if len(pDelivery) < 3 else pDelivery[2]

                    pUrl = item.find_element(By.CLASS_NAME, 'c-card-item__anchor').get_attribute("href")
                    pImgUrl = item.find_element(By.CSS_SELECTOR, '.c-lazyload.c-lazyload--ratio_1x1 img').get_attribute("src")

                    split_price = price_query.split('-')

                    if filter_query in pName and int(split_price[0]) <= int(pPrice) <= int(split_price[1]) :
                        data.append([pName, pPrice, pDelivery_date, pDelivery_fee, pUrl, pImgUrl])
                        cursor.execute("insert into market (num, productName, won, url, delivery, deliveryfee, imgUrl, marketName, keyword) values (seq_market.nextval, :1, :2, :3, :4, :5, :6, :7, :8)",
               [pName, pPrice, pUrl, pDelivery_date, pDelivery_fee, pImgUrl, '11번가', search_query]) #쿼리 실행

                        con.commit()
                        print("상품명" + pName)
                        print('--------------------')
                        # 가격 합 및 상품 개수 업데이트
                        total_price += int(pPrice)
                        count += 1

                except Exception as e:
                    print(f"Error processing item: {e}")

            # 다음 페이지로 이동
            next_page_button_xpath = f'//*[@id="section_commonPrd"]/nav/ul/li[{current_page + 1}]/button'
            next_page_button = driver.find_element(By.XPATH, next_page_button_xpath)

            if "disable" not in next_page_button.get_attribute("class"):
                next_page_button.click()
            else:
                print(f"Next page button for page {current_page + 1} is disabled. Exiting loop.")
                break

            # 페이지 카운트 증가
            current_page += 1

        if not data:
            print("No data found. Please check the HTML structure")
        else:
            # 나머지 코드 생략
            pass

        average_price = total_price / count
        print(f"Average Price: {average_price}")

            # Calculate the range for prices within plus or minus 50%
        price_range_min = average_price * 0.5
        price_range_max = average_price * 1.5

            # Filter products within the price range
        filtered_data = [product for product in data if price_range_min <= int(product[1]) <= price_range_max]

            # Save the filtered data to a new Excel file
        wb_filtered = openpyxl.Workbook()
        ws_filtered = wb_filtered.active
        ws_filtered.title = 'Filtered Products'

            # Write header
        ws_filtered.append(['Product Name', 'Price', 'Delivery Date', 'Delivery Fee', 'URL', 'Image URL'])
        wb_filtered = openpyxl.Workbook()
        ws_filtered = wb_filtered.active
        ws_filtered.title = 'Filtered Products'
        ws_filtered.append(['Product Name', 'Price', 'Delivery Date', 'Delivery Fee', 'URL', 'Image URL'])

            # Write data
        for product in filtered_data:
            ws_filtered.append(product)

        # Save the new Excel file
        wb_filtered.save('filtered_11st_product_info.xlsx')
        
    finally:
        input()
        driver.quit()

# search_query 변수 정의
search_query = input('검색할 단어 : ')
filter_query = input('추려낼 단어 : ')
price_query = input('금액 범위(ex 4000-10000) : ')

# 드라이버 초기화 및 함수 호출
driver = initialize_driver()
search_and_scrape(driver, search_query)
